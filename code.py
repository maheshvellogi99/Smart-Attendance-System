import cv2
import openpyxl
import datetime
import time
import signal
import sys
from firebase_config import get_database
from pyzbar import pyzbar  # Add this import for barcode scanning
from face_utils import recognize_face, register_face  # Add face recognition helpers
import pyttsx3  # Add text-to-speech for audio announcements

# Configuration
EXCEL_FILE = 'attendance.xlsx'
SHEET_NAME = 'Attendance'
WEBCAM_ID = 0  # 0 for default webcam, 1 for external, etc.

# Global variables for cleanup
cap = None
wb = None
engine = None

def signal_handler(signum, frame):
    """Handle Ctrl+C gracefully"""
    print("\nCleaning up and exiting...")
    cleanup()
    sys.exit(0)

def cleanup():
    """Clean up resources before exit"""
    global cap, wb, engine
    if cap is not None:
        cap.release()
    cv2.destroyAllWindows()
    if wb is not None:
        try:
            wb.save(EXCEL_FILE)
        except Exception as e:
            print(f"Error saving Excel file: {e}")
    if engine is not None:
        engine.stop()

def speak_message(message):
    """Speak a message using text-to-speech"""
    global engine
    try:
        if engine is None:
            engine = pyttsx3.init()
            engine.setProperty('rate', 150)  # Speed of speech
        engine.say(message)
        engine.runAndWait()
    except Exception as e:
        print(f"Audio error: {e}")

def capture_face_from_webcam(webcam_id):
    global cap
    cap = cv2.VideoCapture(webcam_id)
    if not cap.isOpened():
        print("Error: Could not open webcam.")
        return None
    
    # Set camera properties for better performance
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
    cap.set(cv2.CAP_PROP_FPS, 30)
    
    print("Please position your face in front of the camera. Press 'c' to capture, or 'ESC' to quit.")
    
    face_not_recognized_shown = False
    start_time = time.time()
    frame_count = 0
    last_recognition_time = 0
    current_roll_no = None
    
    while True:
        ret, frame = cap.read()
        if not ret:
            print("Error: Could not read frame.")
            return None
        
        # Only perform face recognition every 10 frames (about 3 times per second)
        frame_count += 1
        current_time = time.time()
        
        if frame_count % 10 == 0 and (current_time - last_recognition_time) > 0.3:
            # Try to recognize face
            roll_no = recognize_face(frame)
            last_recognition_time = current_time
            if roll_no:
                current_roll_no = roll_no
            else:
                current_roll_no = None
        
        # Display text based on current recognition status
        if current_roll_no:
            # Face recognized - show success message
            cv2.putText(frame, f"Face Recognized: {current_roll_no}", 
                       (10, 50), cv2.FONT_HERSHEY_SIMPLEX, 1.2, (0, 255, 0), 3)
            cv2.putText(frame, "Press 'c' to mark attendance", 
                       (10, 100), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 255, 0), 2)
            face_not_recognized_shown = False
        else:
            # Face not recognized - show message after 2 seconds
            if current_time - start_time > 2:
                cv2.putText(frame, "Face Not Recognized", 
                           (10, 50), cv2.FONT_HERSHEY_SIMPLEX, 1.2, (255, 255, 255), 3)
                cv2.putText(frame, "Please scan ID for registration", 
                           (10, 100), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (255, 255, 255), 2)
                if not face_not_recognized_shown:
                    speak_message("Face not recognized. Please scan your ID for registration.")
                    face_not_recognized_shown = True
        
        cv2.imshow('Face Capture', frame)
        key = cv2.waitKey(1) & 0xFF
        if key == ord('c'):
            cv2.destroyWindow('Face Capture')
            cap.release()
            return frame
        elif key == 27:  # ESC key
            cv2.destroyWindow('Face Capture')
            cap.release()
            print("\nQuitting program...")
            cleanup()
            sys.exit(0)

def read_code_from_webcam(webcam_id):
    """Reads both QR code and barcode data from the webcam feed with visual feedback."""
    global cap
    try:
        cap = cv2.VideoCapture(webcam_id)
        if not cap.isOpened():
            print("Error: Could not open webcam.")
            return None

        # Set camera properties for better performance
        cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
        cap.set(cv2.CAP_PROP_FPS, 30)

        # Initialize QR Code detector
        qr_detector = cv2.QRCodeDetector()
        
        frame_count = 0

        while True:
            ret, frame = cap.read()
            if not ret:
                print("Error: Could not read frame.")
                return None

            # Draw a rectangle in the center of the frame
            height, width = frame.shape[:2]
            center_x, center_y = width // 2, height // 2
            rect_size = 200
            cv2.rectangle(frame, 
                         (center_x - rect_size//2, center_y - rect_size//2),
                         (center_x + rect_size//2, center_y + rect_size//2),
                         (0, 255, 0), 2)

            # Add instruction text
            cv2.putText(frame, "Position QR code or barcode in the green box",
                       (10, 50), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 255, 0), 2)
            cv2.putText(frame, "Press 'q', 'k', or 'ESC' to quit",
                       (10, 100), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 255, 0), 2)

            # Show the frame
            cv2.imshow('Code Scanner', frame)

            # Only scan for codes every 5 frames to reduce lag
            frame_count += 1
            if frame_count % 5 == 0:
                # Try to detect QR code
                retval, decoded_info, points, _ = qr_detector.detectAndDecodeMulti(frame)
                
                if retval and len(decoded_info) > 0 and decoded_info[0]:
                    qr_data = decoded_info[0]
                    # Close the window and release camera before returning
                    cv2.destroyWindow('Code Scanner')
                    cap.release()
                    return qr_data

                # Try to detect barcode
                barcodes = pyzbar.decode(frame)
                for barcode in barcodes:
                    barcode_data = barcode.data.decode('utf-8')
                    # Close the window and release camera before returning
                    cv2.destroyWindow('Code Scanner')
                    cap.release()
                    return barcode_data

            # Check for quit keys
            key = cv2.waitKey(1) & 0xFF
            if key in [ord('q'), ord('k'), 27]:  # 27 is ESC key
                cv2.destroyWindow('Code Scanner')
                cap.release()
                print("\nQuitting program...")
                cleanup()
                sys.exit(0)

        return None

    except Exception as e:
        print(f"Error reading from webcam: {e}")
        if cap is not None:
            cap.release()
        cv2.destroyWindow('Code Scanner')
        return None

def update_cloud_storage(barcode, date_str, time_str):
    """Updates attendance in Firebase Realtime Database with more details."""
    try:
        db = get_database()
        attendance_ref = db.child('attendance')
        
        # Create attendance record with more details
        attendance_data = {
            'time': time_str,
            'status': 'present',
            'last_updated': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        # Update both date-wise and student-wise records
        attendance_ref.child(date_str).child(barcode).set(attendance_data)
        attendance_ref.child('students').child(barcode).child(date_str).set(attendance_data)
        
        print(f"Updated cloud storage for {barcode}")
    except Exception as e:
        print(f"Error updating cloud storage: {e}")
        # Log the error for debugging
        with open('error_log.txt', 'a') as f:
            f.write(f"{datetime.datetime.now()}: Error updating cloud storage for {barcode}: {str(e)}\n")

def initialize_excel_sheet(sheet):
    """Initialize the Excel sheet with proper headers and formatting."""
    # Define headers
    headers = {
        'A1': 'Student ID',
        'B1': 'Year',
        'C1': 'Registration Date'
    }
    
    # Add headers
    for cell, value in headers.items():
        sheet[cell] = value
        sheet[cell].font = openpyxl.styles.Font(bold=True)
        sheet[cell].fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Set column widths
    column_widths = {
        'A': 15,  # Student ID
        'B': 10,  # Year
        'C': 20   # Registration Date
    }
    
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

def process_attendance(barcode, wb, sheet):
    """Processes attendance and updates both Excel and cloud storage with more details."""
    if not barcode:
        return

    now = datetime.datetime.now()
    date_str = now.strftime('%Y-%m-%d')
    time_str = now.strftime('%H:%M:%S')

    try:
        # Check if the date column exists, if not, create it
        header_row = 1
        date_column = None
        for col_num in range(1, sheet.max_column + 1):
            if sheet.cell(row=header_row, column=col_num).value == date_str:
                date_column = col_num
                break

        if not date_column:
            date_column = sheet.max_column + 1
            sheet.cell(row=header_row, column=date_column).value = date_str
            # Format the new date column header
            sheet.cell(row=header_row, column=date_column).font = openpyxl.styles.Font(bold=True)
            sheet.cell(row=header_row, column=date_column).fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            sheet.column_dimensions[openpyxl.utils.get_column_letter(date_column)].width = 15

        # Find the student's row
        student_row = None
        for row_num in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_num, column=1).value == barcode:
                student_row = row_num
                break

        if not student_row:
            # Student not found, add a new row with default values
            student_row = sheet.max_row + 1
            sheet.cell(row=student_row, column=1).value = barcode
            
            # Calculate year as 5 minus 3rd character
            if len(barcode) >= 3:
                try:
                    third_char = int(barcode[2])
                    year = 5 - third_char
                    sheet.cell(row=student_row, column=2).value = str(year)
                except ValueError:
                    sheet.cell(row=student_row, column=2).value = "Not Specified"
            else:
                sheet.cell(row=student_row, column=2).value = "Not Specified"
                
            sheet.cell(row=student_row, column=3).value = now.strftime('%Y-%m-%d %H:%M:%S')
        else:
            # Update year for existing student
            if len(barcode) >= 3:
                try:
                    third_char = int(barcode[1])
                    year = 5 - third_char
                    sheet.cell(row=student_row, column=2).value = str(year)
                except ValueError:
                    pass  # Keep existing year value if 3rd character is not a number

        # Mark attendance with more details
        sheet.cell(row=student_row, column=date_column).value = time_str
        
        print(f"Attendance marked for {barcode} at {time_str}")
        
        # Audio notification for successful attendance marking
        speak_message(f"Attendance marked successfully for {barcode}")

        # Update cloud storage
        update_cloud_storage(barcode, date_str, time_str)

        # Save Excel file with error handling
        try:
            wb.save(EXCEL_FILE)
        except PermissionError:
            print("Error: Excel file is open. Please close it and try again.")
            # Log the error
            with open('error_log.txt', 'a') as f:
                f.write(f"{datetime.datetime.now()}: Excel file permission error for {barcode}\n")
        except Exception as e:
            print(f"Error saving Excel file: {e}")
            # Log the error
            with open('error_log.txt', 'a') as f:
                f.write(f"{datetime.datetime.now()}: Excel save error for {barcode}: {str(e)}\n")
                
    except Exception as e:
        print(f"Error processing attendance: {e}")
        # Log the error
        with open('error_log.txt', 'a') as f:
            f.write(f"{datetime.datetime.now()}: Attendance processing error for {barcode}: {str(e)}\n")

def main():
    """Main function to run the attendance system."""
    global wb
    try:
        # Set up signal handler for Ctrl+C
        signal.signal(signal.SIGINT, signal_handler)
        
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            wb.save(EXCEL_FILE)

        sheet = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.create_sheet(SHEET_NAME)
        
        # Initialize the sheet with proper structure
        initialize_excel_sheet(sheet)

        print("Attendance system started. Scan face or QR codes/barcodes using webcam.")
        print("Press 'c' to capture face, 'q', 'k', or 'ESC' to quit.")
        print("Press Ctrl+C to exit gracefully.")
        
        while True:
            # Step 1: Try to recognize face
            print("Position your face for attendance. Press 'c' to capture, or 'q' to quit.")
            face_frame = capture_face_from_webcam(WEBCAM_ID)
            if face_frame is not None:
                roll_no = recognize_face(face_frame)
                if roll_no:
                    print(f"Face recognized! Roll No: {roll_no}")
                    process_attendance(roll_no, wb, sheet)
                else:
                    print("Face not recognized. Please scan your barcode/QR code to register.")
                    code_data = read_code_from_webcam(WEBCAM_ID)
                    if code_data:
                        # Register the new face
                        success, msg = register_face(face_frame, code_data)
                        print(msg)
                        if success:
                            process_attendance(code_data, wb, sheet)
            else:
                print("No face captured. Try again.")
            time.sleep(1)
            
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        cleanup()

if __name__ == "__main__":
    main()
